"""XLSX parser using openpyxl (already present in dependencies)."""

import time
import logging
from ..parsers import BaseParser, DocumentParseResult, register_parser

logger = logging.getLogger(__name__)


class XLSXParser(BaseParser):
    extension = ".xlsx"
    display_name = "XLSX Parser"

    def _import_library(self):
        import openpyxl  # noqa: F401

    def extract_metadata(self, filepath: str) -> dict:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        sheet_names = wb.sheetnames
        total_rows = 0
        total_cols = 0
        for name in sheet_names:
            ws = wb[name]
            total_rows += ws.max_row or 0
            total_cols = max(total_cols, ws.max_column or 0)
        wb.close()
        return {
            "sheets": len(sheet_names),
            "sheet_names": sheet_names,
            "total_rows": total_rows,
            "total_columns": total_cols,
        }

    def parse(self, filepath: str) -> DocumentParseResult:
        start = time.time()
        try:
            import openpyxl
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            text_parts = []
            sheet_count = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_count += 1
                text_parts.append(f"--- Sheet: {sheet_name} ---")
                rows_added = 0

                for row in ws.iter_rows(values_only=True):
                    cells = [str(cell) if cell is not None else "" for cell in row]
                    line = " | ".join(cells).strip()
                    if line:
                        text_parts.append(line)
                        rows_added += 1

                if rows_added > 0:
                    text_parts.append("")  # blank line between sheets

            wb.close()
            full_text = "\n".join(text_parts)
            word_count = len(full_text.split())
            elapsed = (time.time() - start) * 1000

            logger.debug(
                "Parsed XLSX: %s (%d sheets, %d words, %.0fms)",
                filepath, sheet_count, word_count, elapsed,
            )
            return DocumentParseResult(
                text=full_text,
                metadata=self.extract_metadata(filepath),
                page_count=sheet_count,
                word_count=word_count,
                warnings=[],
                parser_used="xlsx_parser",
                processing_time_ms=round(elapsed, 2),
            )

        except Exception as e:
            elapsed = (time.time() - start) * 1000
            logger.error("XLSX parsing failed for %s: %s", filepath, e)
            return DocumentParseResult(
                text="",
                metadata={},
                page_count=0,
                word_count=0,
                warnings=[f"XLSX parsing failed: {e}"],
                parser_used="xlsx_parser",
                processing_time_ms=round(elapsed, 2),
            )


register_parser(XLSXParser())